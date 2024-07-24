import face_recognition
import cv2
import numpy as np
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import subprocess
import os
import time
import shutil

# Function to load and convert images using face_recognition
def load_and_convert_image(image_path):
    image = face_recognition.load_image_file(image_path)
    return image

# Replace this with your actual image paths
alakh_image_path = "C:\\Users\\uday\\Documents\\Python Scripts\\facerecognization\\pawankalyan.jpg"
ishan_image_path = "C:\\Users\\uday\\Documents\\Python Scripts\\facerecognization\\kajal.jpg"
varun_image_path = "C:\\Users\\uday\\Documents\\Python Scripts\\facerecognization\\mahesh.jpg"

# Load known faces
alakh_image = load_and_convert_image(alakh_image_path)
alakh_encoding = face_recognition.face_encodings(alakh_image)[0]

ishan_image = load_and_convert_image(ishan_image_path)
ishan_encoding = face_recognition.face_encodings(ishan_image)[0]

varun_image = load_and_convert_image(varun_image_path)
varun_encoding = face_recognition.face_encodings(varun_image)[0]

known_face_encodings = [alakh_encoding, ishan_encoding, varun_encoding]
known_face_names = ["Alakh", "Ishan", "Varun"]

# List of expected students
students = known_face_names.copy()

face_locations = []
face_encodings = []

# Get the current date
now = datetime.now()
current_date = now.strftime("%Y-%m-%d")

# Create an Excel workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Attendance"
ws.append(["Name", "Time", "Period"])

# Define the periods
periods = ["Period 1", "Period 2", "Period 3", "Period 4"]

# Function to determine the current period
def get_current_period():
    current_hour = datetime.now().hour
    if 9 <= current_hour < 10:
        return "Period 1"
    elif 10 <= current_hour < 11:
        return "Period 2"
    elif 11 <= current_hour < 12:
        return "Period 3"
    elif 12 <= current_hour < 13:
        return "Period 4"
    else:
        return "Unknown Period"

# Save the Excel workbook initially
excel_path = f"{current_date}.xlsx"
temp_excel_path = f"temp_{current_date}.xlsx"
wb.save(temp_excel_path)

# Start video capture
video_capture = cv2.VideoCapture(0)

while True:
    ret, frame = video_capture.read()
    if not ret:
        break

    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    # Recognize faces
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        face_distance = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_index = np.argmin(face_distance)

        if matches[best_match_index]:
            name = known_face_names[best_match_index]

            # Add the text if a person is present
            if name in known_face_names:
                font = cv2.FONT_HERSHEY_SIMPLEX
                bottomLeftCornerOfText = (left * 4, bottom * 4 + 20)
                fontScale = 1
                fontColor = (255, 0, 0)
                thickness = 2
                lineType = 2
                cv2.putText(frame, name + " Present", bottomLeftCornerOfText, font, fontScale, fontColor, thickness, lineType)

                if name in students:
                    print(f"{name} detected, logging attendance")
                    current_time = datetime.now().strftime("%H:%M:%S")
                    current_period = get_current_period()
                    ws.append([name, current_time, current_period])
                    students.remove(name)  # Ensure each student is logged only once

                    # Save the workbook after every update
                    for attempt in range(5):
                        try:
                            wb.save(temp_excel_path)
                            break
                        except PermissionError:
                            print(f"Attempt {attempt+1}: PermissionError: Could not save {temp_excel_path}. Retrying in 1 second...")
                            time.sleep(1)

    cv2.imshow("Camera", frame)

    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

# Release video capture and close all OpenCV windows
video_capture.release()
cv2.destroyAllWindows()

# Save the final Excel workbook to the final path
for attempt in range(5):
    try:
        wb.save(excel_path)
        break
    except PermissionError:
        print(f"Attempt {attempt+1}: PermissionError: Could not save {excel_path}. Retrying in 1 second...")
        time.sleep(1)

# Open the final Excel file using the default application
if os.name == 'nt':  # For Windows
    subprocess.Popen(["start", excel_path], shell=True)
elif os.name == 'posix':  # For macOS and Linux
    subprocess.Popen(["open", excel_path])

# Open and read the final Excel file to verify contents
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Print the contents of the Excel file
for row in ws.iter_rows(values_only=True):
    print(row)
